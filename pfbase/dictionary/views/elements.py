from rest_framework import status, views, exceptions as exc
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from pfbase.base_views import AbstractModelAPIView
from pfbase.pagination import CustomPagination
from ..serializers.elements import EIGetSerializer, EIPostSerializer, EIUpdateSerializer
from ..models import elements, dictionaries, indicators
from ..models import Elements
from ...system.models import organization
from rest_framework.parsers import MultiPartParser, FormParser
from ..service import find_driver
import openpyxl


class ElementsAPIView(AbstractModelAPIView):
    """
    Представление :Element
    """
    queryset = elements.Elements.objects.all().order_by('-id')
    serializer_class = EIGetSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = CustomPagination

    @action(detail=False, methods=['get'], url_path='bydictionary/(?P<pk>\d+)')
    def bydictionary(self, request, pk=None):
        """
        Получения элементов по :pk ABCDictionary
        """
        dictionary = elements.Dictionaries.objects.get(pk=pk)
        params = request.GET.copy()
        params["DICT_CODE"] = dictionary.code
        indicators = elements.Elements.objects.getByFilter(params)       
        page = self.paginate_queryset(indicators)
        if page is not None:
            serializer = EIGetSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)
        if not indicators:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = EIGetSerializer(indicators, many=True, context={'request': request})
        return Response(serializer.data)


class EIAPIView(views.APIView):
    """
    Представление :Element with their :Indicator
    """
    queryset = elements.Elements.objects.all()
    permission_classes = (IsAuthenticated,)

    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk', False)
        if not pk:
            query_params = request.query_params
            code = kwargs.get('code', False)
            elements_queryset = self.queryset.filter(dictionary__code=code)
            instance = find_driver(elements_queryset, query_params)
            return Response(instance, status=status.HTTP_200_OK)
        try:
            queryset = self.queryset.get(id=pk)
        except elements.Elements.DoesNotExist:
            return Response({"message": "Element not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(EIGetSerializer(queryset, many=False, context={'request': request}).data)

    def post(self, request):
        serializer = EIPostSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Element created successfully"},
                            status=status.HTTP_201_CREATED)
        raise exc.ValidationError(serializer.errors)

    def patch(self, request, *args, **kwargs):
        serializer = EIUpdateSerializer(data=request.data, context={'request': request})
        pk = kwargs.get('pk', False)
        if not pk:
            return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        try:
            element = self.queryset.get(pk=pk)
        except elements.Elements.DoesNotExist:
            return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        if serializer.is_valid():
            serializer.update(element, serializer.validated_data)
            return Response({"message": "Element updated successfully"})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, *args, **kwargs):
        pk = kwargs.get('pk', False)
        if not pk:
            return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        try:
            element = self.queryset.get(pk=pk)
            element.delete()
        except elements.Elements.DoesNotExist:
            return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class FileUploadView(views.APIView):
    parser_classes = (MultiPartParser, FormParser)
    def post(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            code = uploaded_file.name.split('_', 1)[-1].split('.')[0].strip()
            if not dictionaries.Dictionaries.objects.filter(code=code).exists():
                return Response({"error": f"Code '{code}' not found in the database"},
                                status=status.HTTP_404_NOT_FOUND)
            dictionary_id = dictionaries.Dictionaries.objects.getByCode(code)
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            if 'PF' not in workbook.sheetnames:
                return Response({"error": "Sheet 'PF' not found"}, status=status.HTTP_400_BAD_REQUEST)
            sheet = workbook['PF']
            headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if
                       cell is not None]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(row)
                row_data = dict(zip(headers, row))
                if all(value is None for value in row_data.values()):
                    break

                parent_id = None
                element = None #for updating
                organization_id = None
                indicators_list = []
                short_name = {}
                code_el = None

                if 'PARENT.CODE' in headers:
                    parent_code = row_data.get("PARENT.CODE")
                    if parent_code:
                        try:
                            parent_element = Elements.objects.get(code=str(parent_code).strip())
                            if parent_element:
                                parent_id = parent_element.id
                        except:
                            continue
                for header, value in row_data.items():
                    if header.startswith("SHORTNAME.") and value:
                        lang_code = header.split(".")[1]
                        short_name[lang_code] = value
                
                raw_code_el = row_data.get("CODE(pk)")
                
                if raw_code_el:
                    
                    code_el = str(raw_code_el).strip()
                    
                    try:
                        print(f"Checking element with code: '{code_el}'")

                        check_element = Elements.objects.filter(code=code_el).first()
                        if check_element:
                            # if Element exists -> update
                            element = check_element

                            code_el = None 
                    except Elements.DoesNotExist:
                        code_el = row_data.get("CODE(pk)").strip()
                        

                if code_el is None and not element:
                    # If neither creating nor updating, skip
                    continue

                if "ORGANIZATION.CODE" in header:
                    org_code = row_data.get("ORGANIZATION.CODE").strip()            
                    if org_code:
                        organization_id = organization.Organization.objects.getByCode(org_code)
                for header, value in row_data.items():
                    if header and header.startswith("IDC.") and value is not None:
                        parts = header.split('.')
                        if len(parts) >= 3:
                            ind_code = parts[1]
                            data_type = parts[2]
                            ind_type = {
                                "int": "int", "str": "str", "date": "date",
                                "time": "time", "bool": "bool", "float": "float",
                                "datetime": "datetime", "text": "text", "dct": "dct"
                            }.get(data_type)
                            if not ind_type:
                                continue
                            indicator_entry = indicators.DctIndicators.objects.getByCode(ind_code)
                            if ind_type == "dct" or ind_type == "dcm" or ind_type == "list":
                                try:
                                    element_id = Elements.objects.getByCode(value)
                                    if element_id:
                                        value = element_id
                                except Elements.DoesNotExist:
                                    continue
                            if ind_type == "date":
                                value = value.strftime("%Y-%m-%d")
                                
                            if indicator_entry:
                                indicators_list.append({
                                    "id": indicator_entry,
                                    "code": ind_code,
                                    "value": value,
                                    "type": ind_type,
                                })

                json_entry = {
                    "short_name": short_name,
                    "dictionary_id": dictionary_id,
                    "indicators": indicators_list,
                }
                
                if code_el is not None:
                    json_entry["code"] = code_el
                if organization_id is not None:
                    json_entry["organization_id"] = organization_id
                if parent_id is not None:
                    json_entry["parent_id"] = parent_id

                #comment element if trying to update
                element = None
                #
                if element and code_el is None:
                    serializer = EIUpdateSerializer(element, data=json_entry, partial=True,
                                                    context={'request': request})
                    if serializer.is_valid():
                        serializer.save()
                    else:
                        return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                elif code_el and not element:
                    serializer = EIPostSerializer(data=json_entry, context={'request': request})
                    if serializer.is_valid():
                        serializer.save()
                    else:
                        return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            return Response({"message": "All elements processed successfully"}, status=status.HTTP_201_CREATED)

        except openpyxl.utils.exceptions.InvalidFileException:
            return Response({"error": "Invalid file format. Please upload a valid Excel file."},
                            status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)