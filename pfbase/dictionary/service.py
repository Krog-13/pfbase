import json
from django.db import transaction
from collections import namedtuple
from pfbase.exception import WrongType
from django.utils import timezone
from datetime import datetime
from ..system import models as stm_models
from .models import Elements, DctIndicators, Dictionaries
from ..document import models as dcm_models
from rest_framework import status
from rest_framework.response import Response
from .models import dictionaries, indicators
from ..system.models import organization, ListValues, Organization
import openpyxl
import mimetypes
from pfbase.exception import ExcelFormatError


Typing = namedtuple('Typing', ['int', 'float', 'str', 'text', 'datetime', 'bool', 'reference', 'json', 'file', 'user', 'array_int', 'array_str'])
marker = Typing(int="int", float="float", str="str", text="text", json='json', file="file", user="user",
                datetime=["datetime", "date", "time"], bool="bool", reference=["dct", "list", "dcm", "user", "org"], array_int=["dct", "dcm", "list", "user", "org"],
                array_str=["str", "text", "bool", "file", "date", "time", "dct"])


class ElementService:

    def __init__(self):
        self.short_name = None
        self.full_name = None
        self.code = None
        self.dictionary_id = None
        self.parent_id = None
        self.indicators = []

    def validate_data(self, request_data):
        """Validation for creating a new element"""
        self.short_name = request_data.get('short_name')
        self.full_name = request_data.get('full_name')
        self.code = request_data.get('code')
        self.dct_code = request_data.get('dct_code')
        self.dictionary_id = request_data.get('dictionary_id')
        self.dictionary_code = request_data.get('dictionary_code')
        self.parent_id = request_data.get('parent_id')
        self.parent_code = request_data.get('parent_code')
        self.organization_id = request_data.get('organization_id')
        self.organization_code = request_data.get('organization_code')
        self.indicators = request_data.get('indicators', [])

        if not self.dictionary_id and not self.code:
            raise WrongType("Invalid data: 'dictionary_id' or 'code' is required.")

    def validate_update_data(self, request_data):
        """Validation for updating an existing element"""
        self.short_name = request_data.get('short_name')
        self.full_name = request_data.get('full_name')
        self.code = request_data.get('code')
        self.dictionary_id = request_data.get('dictionary_id')
        self.dictionary_code = request_data.get('dictionary_code')
        self.parent_id = request_data.get('parent_id')
        self.parent_code = request_data.get('parent_code')
        self.organization_id = request_data.get('organization_id')
        self.organization_code = request_data.get('organization_code')
        self.indicators = request_data.get('indicators', [])

    @transaction.atomic
    def create_element_iv(self, user, validated_data):
        self.validate_data(validated_data)
        dictionary = Dictionaries.objects.get(id=self.dictionary_id) if self.dictionary_id else Dictionaries.objects.get(code=self.dictionary_code)
        if self.parent_code:
            parent_e = Elements.objects.get(code=self.parent_code)
        elif self.parent_id:
            parent_e = Elements.objects.get(id=self.parent_id)
        else:
            parent_e = None

        org = Organization.objects.get(id=self.organization_id) if self.organization_id else Organization.objects.get(code=self.organization_code)

        element = Elements.objects.create(
            short_name=self.short_name,
            full_name=self.full_name,
            code=self.code,
            dictionary=dictionary,
            author=user,
            parent=parent_e,
            organization=org
        )

        if not self.indicators:
            return element

        for indicator in self.indicators:
            self.create_or_update_indicator_value(element, indicator)

        return element

    @transaction.atomic
    def create_list_element(self, user, validated_data_list):
        for validated_data in validated_data_list["elements"]:
            self.validate_data(validated_data)
            dictionary = Dictionaries.objects.get(
                id=self.dictionary_id) if self.dictionary_id else Dictionaries.objects.get(code=self.dct_code)
            parent_e = Elements.objects.get(id=self.parent_id) if self.parent_id else None

            element = Elements.objects.create(
                short_name=self.short_name,
                full_name=self.full_name,
                code=self.code,
                dictionary=dictionary,
                author=user,
                parent=parent_e,
                organization_id=self.organization_id
            )

            if not self.indicators:
                continue

            for indicator in self.indicators:
                self.create_or_update_indicator_value(element, indicator)
        return {}

    @transaction.atomic
    def update_element_iv(self, element, user, validated_data):
        self.validate_update_data(validated_data)
        if self.parent_id:
            element.parent = Elements.objects.get(id=self.parent_id)
        if self.short_name:
            element.short_name = self.short_name
        if self.full_name:
            element.full_name = self.full_name
        if self.code:
            element.code = self.code
        if self.organization_id:
            element.organization_id = self.organization_id

        element.author = user
        element.save()

        for indicator in self.indicators:
            self.create_or_update_indicator_value(element, indicator, update=True)

        return element

    def create_or_update_indicator_value(self, element, indicator, update=False):
        some_value = indicator.get('value')
        type_value = indicator.get('type')
        idc_id = indicator.get('id')
        idc_code = indicator.get('code')

        if type_value == marker.reference[1]:
            if not isinstance(some_value, int):
                if not some_value.isdigit():
                    some_value = stm_models.ListValues.objects.get(code=some_value).id
                else:
                    stm_models.ListValues.objects.get(id=some_value)
        elif type_value == marker.reference[0]:
            if not isinstance(some_value, int):
                if not some_value.isdigit():
                    some_value = Elements.objects.get(code=some_value).id
                else:
                    Elements.objects.get(id=some_value)
        elif type_value == marker.reference[2]:
            if not isinstance(some_value, int):
                if not some_value.isdigit():
                    some_value = dcm_models.Records.objects.get(code=some_value).id
                else:
                    dcm_models.Records.objects.get(id=some_value)

        if idc_id:
            dct_indicator = DctIndicators.objects.get(id=idc_id, type_value=type_value)
        else:
            dct_indicator = DctIndicators.objects.get(code=idc_code, type_value=type_value)

        ev = element.element_values.filter(indicator=dct_indicator).first() if update else None
        if ev is None:
            ev = element.element_values.create(indicator=dct_indicator)

        if not self.separate_value(ev, type_value, some_value):
            raise WrongType("Invalid type value")
        ev.save()

    def separate_value(self, entity, type_value, some_value):
        if type_value == marker.int:
            entity.value_int = some_value
        elif type_value == marker.float:
            entity.value_float = some_value
        elif type_value == marker.str:
            entity.value_str = some_value
        elif type_value == marker.text:
            entity.value_text = some_value
        elif type_value == marker.bool:
            entity.value_bool = some_value
        elif type_value in marker.reference:
            entity.value_reference = some_value
        elif type_value == marker.json:
            entity.value_json = some_value
        elif type_value in marker.datetime:
            date = self.give_date_format(some_value, type_value)
            if not date:
                return False
            entity.value_datetime = date
        else:
            return False
        return True

    @staticmethod
    def give_date_format(date_str, type_value):
        """
        Convert str: datetime, date, time to datetime object
        """
        formats = [
            ("%Y-%m-%d %H:%M", "datetime"),  # "2024-01-01 12:33"
            ("%Y-%m-%d", "date"),  # "2024-01-01"
            ("%H:%M", "time")  # "10:23"
        ]
        try:
            for fmt in formats:
                if type_value == fmt[1]:
                    parse_datetime = datetime.strptime(date_str, fmt[0])
                    if type_value == "time":
                        parse_datetime = datetime.combine(datetime.now().date(), parse_datetime.time())
                    return timezone.make_aware(parse_datetime)
        except ValueError:
            return False

class ExcelUpload:
    """
    Upload excel file Dictionaries
    """
    def __init__(self, excel_file, user, trigger):
        self.excel_file = excel_file
        self.user = user
        self.trigger = trigger
        self.workbook = None
        self.sheet = None
        self.dct_code = None
        self.dct_id = None
        self.dcm_code = None
        self.dcm_id = None
        self.mapping = {}
        self.output = {}
        self.element_service = ElementService()

    def start_upload(self):
        self._check()
        headers = next(self.sheet.iter_rows(max_row=1, values_only=True))
        for idx, item in  enumerate(headers, start=1):
            if item == "CODE":
                self.mapping[idx] = "code"
            elif item == "SHORTNAME.RU":
                self.mapping[idx] = "short_name.ru"
            elif item == "SHORTNAME.KK":
                self.mapping[idx] = "short_name.kk"
            elif item == "SHORTNAME.EN":
                self.mapping[idx] = "short_name.en"
            elif item == "FULLNAME.RU":
                self.mapping[idx] = "full_name.ru"
            elif item == "FULLNAME.KK":
                self.mapping[idx] = "full_name.kk"
            elif item == "FULLNAME.EN":
                self.mapping[idx] = "full_name.en"
            elif item == "PARENT.CODE":
                self.mapping[idx] = "parent_id"
            elif item.startswith("ORGANIZATION"):
                self.mapping[idx] = "organization_id"
            elif item.startswith("IDC"):
                self.mapping[idx] = f"indicators_{item}"

        for row in self.sheet.iter_rows(min_row=2):
            self.output = {"code": None,
                           "dictionary_code": self.dct_code,
                           "parent_id": None,
                           "organization_id": None,
                           "short_name": {"ru": "", "kk": "", "en": ""},
                           "full_name": {"ru": "", "kk": "", "en": ""},
                           "indicators": []}
            short_name_default = {"ru": "", "kk": "", "en": ""}
            full_name_default = {"ru": "", "kk": "", "en": ""}
            for cell in row:
                key = self.mapping[cell.col_idx]
                if key.startswith("short_name"):
                    lang = key[-2:]
                    short_name_default[lang] = cell.value if cell.value else ""
                    self.output["short_name"] = short_name_default
                elif key.startswith("full_name"):
                    lang = key[-2:]
                    full_name_default[lang] = cell.value if cell.value else ""
                    self.output["full_name"] = full_name_default
                elif key == "parent_id":
                    if cell.value:
                        parent_id = Elements.objects.getByCode(cell.value)
                        self.output["parent_id"] = parent_id
                elif key == "organization_id":
                    if cell.value:
                        org_id = Organization.objects.getByCode(cell.value)
                        self.output["organization_id"] = org_id
                elif key.startswith("indicators"):
                    idc = key.split(".")
                    idc_type = idc[2].lower()
                    try:
                        converted_value = self._get_optimal_value(idc_type, cell.value)
                    except json.JSONDecodeError:
                        continue
                    row = {"code": idc[1], "value": converted_value, "type": idc_type}
                    self.output["indicators"].append(row)
                else:
                    self.output[key] = cell.value
            if self.trigger == "create":
                self.element_service.create_element_iv(self.user, self.output)
            elif self.trigger == "update":
                element_code = self.output["code"]
                element = Elements.objects.filter(code=element_code).first()
                if not element:
                    continue
                self.element_service.update_element_iv(element, self.user, self.output)
            elif self.trigger == "create_update":
                element_code = self.output["code"]
                element = Elements.objects.filter(code=element_code).first()
                if not element:
                    self.element_service.create_element_iv(self.user, self.output)
                else:
                    self.element_service.update_element_iv(element, self.user, self.output)
        return self.trigger

    def _check(self):
        if self.excel_file.name.endswith(".xlsx"):
            self.dct_code = self.excel_file.name[:-5].split("_", 1)[1]
            print(self.dct_code)

        self.workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        if self.dct_code and self.workbook.active.title == "PF":
            self.dct_id = Dictionaries.objects.getByCode(code=self.dct_code)
            self.sheet = self.workbook["PF"]
            return
        raise

    def _get_optimal_value(self, type_value, value):
        """
        Comparison type of value and get value
        """
        if not value:
            return None
        if type_value == marker.int:
            return int(value)
        elif type_value == marker.float:
            return float(value)
        elif type_value == marker.str:
            return value
        elif type_value == marker.file:
            return value
        elif type_value == marker.user:
            return value
        elif type_value == marker.text:
            return value
        elif type_value == marker.bool:
            if value.lower() == "false":
                return False
            elif value.lower() == "true":
                return True
            return False
        elif type_value in marker.reference:
            return self.get_id_by_code(type_value, value)
        elif type_value == marker.json:
            return json.loads(value)
        elif type_value in marker.datetime:
            return value
        else:
            return None

    @staticmethod
    def get_id_by_code(type_value, code):
        if type_value == "dct":
            return Elements.objects.getByCode(code=code)
        elif type_value == "list":
            return ListValues.objects.getByCode(code=code)

    @staticmethod
    def give_date_format(date_str, type_value):
        """
        Convert str: datetime, date, time to datetime object
        """
        formats = [
            ("%Y-%m-%d %H:%M", "datetime"),  # "2024-01-01 12:33"
            ("%Y-%m-%d", "date"),  # "2024-01-01"
            ("%H:%M", "time")  # "10:23"
        ]
        try:
            for fmt in formats:
                if type_value == fmt[1]:
                    parse_datetime = datetime.strptime(date_str, fmt[0])
                    if type_value == "time":
                        parse_datetime = datetime.combine(datetime.now().date(), parse_datetime.time())
                    return timezone.make_aware(parse_datetime)
        except ValueError:
            return False

    @staticmethod
    def check_format(file_object):
        """Check file MIME type"""
        # Get the MIME type
        mime_type, _ = mimetypes.guess_type(file_object.name)
        # Allowed MIME types for Excel files
        allowed_mime_types = [
            "application/vnd.ms-excel",  # .xls
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        ]
        if mime_type not in allowed_mime_types:
            raise ExcelFormatError


def find_driver(queryset, params):
    output = {"drivers": []}
    indicators = params.get('indicators')
    drivers = queryset.filter(element_values__value_str__contains=params.get('search'))
    for driver in drivers:
        driver_iv = driver.element_values.all()
        values = {"data": [], "id": None}
        for idc in indicators.split(","):
            value = driver_iv.filter(indicator__code=idc).first()
            if value:
                values["data"].append(value.value_str)
        values["id"] = driver.id
        output["drivers"].append(values)
    return output

def upload_file(uploaded_file):
        json_entry_list = []
        update_json_entry_list = []
        if not uploaded_file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            code = uploaded_file.name.split('_', 1)[-1].split('.')[0].strip()
            if not dictionaries.Dictionaries.objects.filter(code=code).exists():
                raise Dictionaries.objects.DoesNotExist(f"Dictionary with code {code} not found")
            dictionary_id = dictionaries.Dictionaries.objects.getByCode(code)
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            if 'PF' not in workbook.sheetnames:
                raise Exception("Sheet 'PF' not found in the file")
            
            sheet = workbook['PF']
            headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if
                       cell is not None]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))#коннект каждой строки с хедером
                if all(value is None for value in row_data.values()):
                    break

                parent_id = None
                element = None #for updating
                organization_id = None
                indicators_list = []
                short_name = {'kk': '', 'en': '', 'ru': ''}
                code_el = None

                if 'PARENT.CODE' in headers:
                    parent_code = row_data.get("PARENT.CODE")
                    if parent_code:
                        try:
                            parent_element = Elements.objects.get(code=str(parent_code).strip())
                            if parent_element:
                                parent_id = parent_element.id
                        except:
                            continue
                if "ORGANIZATION.CODE" in headers:
                    org_code = row_data.get("ORGANIZATION.CODE").strip()            
                    if org_code:
                        organization_id = organization.Organization.objects.getByCode(org_code)

                for header, value in row_data.items():
                    if header.startswith("SHORTNAME.") and value:
                        lang_code = header.split(".")[1]
                        short_name[lang_code] = value
                        if short_name['kk'] is None:
                            short_name['kk'] = ""
                        if short_name['en'] is None:
                            short_name['en'] = ""
                        if short_name['ru'] is None:
                            short_name['ru'] = ""
                
                raw_code_el = row_data.get("CODE(pk)")
                if raw_code_el:
                    code_el = str(raw_code_el).strip()             
                    try:
                        check_element = Elements.objects.filter(code=code_el).first()
                        if check_element:
                            # if Element exists -> update
                            element = check_element
                            code_el = None 
                    except Elements.DoesNotExist:
                        code_el = row_data.get("CODE(pk)").strip()
                        
                if code_el is None and not element:
                    continue

                for header, value in row_data.items():
                    if header and header.startswith("IDC.") and value is not None:
                        parts = header.split('.')
                        if len(parts) >= 3:
                            ind_code = parts[1]
                            ind_type = parts[2]
                            if not ind_type:
                                continue
                            indicator_entry = indicators.DctIndicators.objects.getByCode(ind_code)
                            if ind_type == "dct" or ind_type == "dcm":
                                try:
                                    element_id = Elements.objects.getByCode(value)
                                    if element_id:
                                        value = element_id
                                except Elements.DoesNotExist:
                                    continue
                            if ind_type == "list":
                                try:
                                    list_id = ListValues.objects.getByCode(value)
                                    if list_id:
                                        value = list_id
                                except ListValues.DoesNotExist:
                                    continue
                            if ind_type == "date":
                                value = value.strftime("%Y-%m-%d")
                            if ind_type == "json":
                                value = json.loads(value)
                            if indicator_entry:
                                indicators_list.append({
                                    "id": indicator_entry,
                                    "code": ind_code,
                                    "value": value,
                                    "type": ind_type,
                                })
                if code_el is not None and element is None:
                    json_entry = {
                        "short_name": short_name,
                        "dictionary_id": dictionary_id,
                        "indicators": indicators_list,
                        "code": code_el,
                    }
                    if organization_id is not None:
                        json_entry["organization_id"] = organization_id
                    if parent_id is not None:
                        json_entry["parent_id"] = parent_id
                    json_entry_list.append(json_entry)

                if element is not None and code_el is None:
                    json_entry = {
                        "short_name": short_name,
                        "indicators": indicators_list,
                        "dictionary_id": dictionary_id,
                    }
                    if parent_id is not None:
                        json_entry["parent_id"] = parent_id
                    if organization_id is not None:
                        json_entry["organization_id"] = organization_id
                    update_json_entry_list.append((element, json_entry))
            return json_entry_list,update_json_entry_list
        except openpyxl.utils.exceptions.InvalidFileException:
            raise Exception("Invalid file format")
        except Exception as e:
            raise Exception(str(e))
